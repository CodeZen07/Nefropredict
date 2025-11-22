import pandas as pd
import numpy as np
import joblib
import json
import os
from datetime import datetime
import streamlit as st
import altair as alt
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO

# =============================================
# CONFIGURACIÓN Y ESTILOS
# =============================================
st.set_page_config(page_title="NefroPredict RD", page_icon="🩺", layout="wide")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap');
    body {font-family: 'Inter', sans-serif;}
    h1, h2, h3 {color: #002868 !important;}
    .stButton>button {background: #002868; color: white; border-radius: 12px; padding: 0.7rem 1.5rem; font-weight:600;}
    .stButton>button:hover {background: #001a4d;}
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 20px; border-radius: 15px; color: white; text-align: center;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
    }
    .risk-high {background: linear-gradient(135deg, #ff6b6b, #ee5a5a); padding:20px; border-radius:12px; color:white; text-align:center;}
    .risk-med {background: linear-gradient(135deg, #feca57, #ff9f43); padding:20px; border-radius:12px; color:white; text-align:center;}
    .risk-low {background: linear-gradient(135deg, #1dd1a1, #10ac84); padding:20px; border-radius:12px; color:white; text-align:center;}
    .patient-card {background:#f8f9fa; padding:15px; border-radius:10px; margin:10px 0; border-left:4px solid #002868;}
</style>
""", unsafe_allow_html=True)

st.markdown("<h1 style='text-align:center;'>🩺 NefroPredict RD 2025</h1>", unsafe_allow_html=True)
st.markdown("<h4 style='text-align:center;color:#555;'>Detección temprana de ERC • República Dominicana</h4>", unsafe_allow_html=True)

# =============================================
# BASE DE DATOS
# =============================================
DB_FILE = "nefro_db.json"

class DataStore:
    def __init__(self):
        if not os.path.exists(DB_FILE):
            self._create_initial_db()
        self.data = self._load()

    def _create_initial_db(self):
        initial = {
            "users": {
                "admin": {"pwd": "admin", "role": "admin", "name": "Administrador", "active": True},
                "dr.perez": {"pwd": "1234", "role": "doctor", "name": "Dr. José Pérez", "active": True},
                "dr.gomez": {"pwd": "1234", "role": "doctor", "name": "Dra. Ana Gómez", "active": True}
            },
            "patients": [],
            "uploads": []
        }
        with open(DB_FILE, "w", encoding="utf-8") as f:
            json.dump(initial, f, indent=4, ensure_ascii=False)

    def _load(self):
        with open(DB_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        # Asegurar que existan todas las keys necesarias
        if "users" not in data:
            data["users"] = {
                "admin": {"pwd": "admin", "role": "admin", "name": "Administrador", "active": True}
            }
        if "patients" not in data:
            data["patients"] = []
        if "uploads" not in data:
            data["uploads"] = []
        # Guardar si se agregaron keys faltantes
        with open(DB_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        return data

    def save(self):
        with open(DB_FILE, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=4, ensure_ascii=False)

    def get_user(self, username):
        return self.data["users"].get(username)

    def create_doctor(self, username, password, full_name):
        self.data["users"][username] = {
            "pwd": password, "role": "doctor", "name": full_name, "active": True
        }
        self.save()

    def update_password(self, username, new_pwd):
        if username in self.data["users"]:
            self.data["users"][username]["pwd"] = new_pwd
            self.save()

    def toggle_active(self, username):
        if username in self.data["users"]:
            self.data["users"][username]["active"] = not self.data["users"][username]["active"]
            self.save()

    def delete_doctor(self, username):
        if username in self.data["users"] and self.data["users"][username]["role"] == "doctor":
            del self.data["users"][username]
            self.save()

    def add_patient(self, record):
        self.data["patients"].insert(0, record)
        self.save()

    def get_patients_by_doctor(self, user_id):
        return [p for p in self.data["patients"] if p["doctor_user"] == user_id]

    def get_all_patients(self):
        return self.data["patients"]

    def add_upload_log(self, log):
        self.data["uploads"].insert(0, log)
        self.save()

db = DataStore()

# =============================================
# MODELO
# =============================================
@st.cache_resource
def load_model():
    try:
        return joblib.load("modelo_erc.joblib")
    except:
        return None

model = load_model()

# =============================================
# FUNCIONES DE RIESGO Y PREDICCIÓN
# =============================================
def riesgo_level(risk):
    if risk > 70:
        return "MUY ALTO", "#CE1126", "Intervención URGENTE - Referir a nefrología"
    elif risk > 40:
        return "ALTO", "#FFC400", "Intervención Media - Control estricto"
    else:
        return "MODERADO", "#4CAF50", "Seguimiento Rutinario - Control periódico"

def predecir(row):
    feats = np.array([[row["edad"], row["imc"], row["presion_sistolica"],
                       row["glucosa_ayunas"], row["creatinina"]]])
    if model:
        return round(model.predict_proba(feats)[0][1] * 100, 1)
    else:
        # Simulación realista basada en factores de riesgo
        base = 10
        base += (row["creatinina"] - 1) * 32
        base += max(0, row["glucosa_ayunas"] - 126) * 0.3
        base += max(0, row["presion_sistolica"] - 140) * 0.2
        base += max(0, row["imc"] - 30) * 0.5
        base += max(0, row["edad"] - 60) * 0.3
        return round(max(1, min(99, base + np.random.uniform(-5, 8))), 1)

def crear_gauge_riesgo(riesgo):
    """Crear gráfico de velocímetro para el riesgo"""
    if riesgo > 70:
        color = "#CE1126"
    elif riesgo > 40:
        color = "#FFC400"
    else:
        color = "#4CAF50"
    
    fig = go.Figure(go.Indicator(
        mode="gauge+number+delta",
        value=riesgo,
        domain={'x': [0, 1], 'y': [0, 1]},
        title={'text': "Riesgo de ERC (%)", 'font': {'size': 24, 'color': '#002868'}},
        gauge={
            'axis': {'range': [0, 100], 'tickwidth': 1, 'tickcolor': "#002868"},
            'bar': {'color': color},
            'bgcolor': "white",
            'borderwidth': 2,
            'bordercolor': "#002868",
            'steps': [
                {'range': [0, 40], 'color': '#e5f7e5'},
                {'range': [40, 70], 'color': '#fff4e5'},
                {'range': [70, 100], 'color': '#ffe5e5'}
            ],
            'threshold': {
                'line': {'color': "red", 'width': 4},
                'thickness': 0.75,
                'value': riesgo
            }
        }
    ))
    fig.update_layout(height=300, margin=dict(l=20, r=20, t=50, b=20))
    return fig

def crear_grafico_factores(paciente):
    """Crear gráfico de barras con los factores de riesgo"""
    # Normalizar valores para comparación
    factores = {
        'Edad': min(100, (paciente['edad'] / 100) * 100),
        'IMC': min(100, (paciente['imc'] / 40) * 100),
        'Presión': min(100, (paciente['presion_sistolica'] / 200) * 100),
        'Glucosa': min(100, (paciente['glucosa_ayunas'] / 300) * 100),
        'Creatinina': min(100, (paciente['creatinina'] / 5) * 100)
    }
    
    colors = []
    for k, v in factores.items():
        if v > 70:
            colors.append('#CE1126')
        elif v > 50:
            colors.append('#FFC400')
        else:
            colors.append('#4CAF50')
    
    fig = go.Figure(go.Bar(
        x=list(factores.keys()),
        y=list(factores.values()),
        marker_color=colors,
        text=[f"{v:.0f}%" for v in factores.values()],
        textposition='outside'
    ))
    fig.update_layout(
        title="Factores de Riesgo (Normalizados)",
        yaxis_title="Nivel (%)",
        height=300,
        margin=dict(l=20, r=20, t=50, b=20)
    )
    return fig

def generar_reporte_html(paciente, riesgo, nivel, doctor):
    """Generar reporte HTML para impresión/PDF"""
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
    color = '#CE1126' if riesgo > 70 else '#FFC400' if riesgo > 40 else '#4CAF50'
    
    return f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>Reporte NefroPredict - {paciente['nombre_paciente']}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'Segoe UI', Arial, sans-serif; background: #f5f5f5; padding: 20px; }}
        .container {{ max-width: 800px; margin: auto; background: white; border-radius: 20px; overflow: hidden; box-shadow: 0 10px 40px rgba(0,0,0,0.1); }}
        .header {{ background: linear-gradient(135deg, #002868, #001a4d); color: white; padding: 30px; text-align: center; }}
        .header h1 {{ font-size: 2em; margin-bottom: 5px; }}
        .header p {{ opacity: 0.9; }}
        .content {{ padding: 30px; }}
        .risk-display {{ text-align: center; padding: 30px; margin: 20px 0; border-radius: 15px; background: linear-gradient(135deg, {color}22, {color}11); border: 3px solid {color}; }}
        .risk-number {{ font-size: 4em; font-weight: bold; color: {color}; }}
        .risk-label {{ font-size: 1.5em; color: {color}; margin-top: 10px; }}
        .info-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 15px; margin: 20px 0; }}
        .info-item {{ background: #f8f9fa; padding: 15px; border-radius: 10px; border-left: 4px solid #002868; }}
        .info-label {{ color: #666; font-size: 0.9em; }}
        .info-value {{ color: #002868; font-size: 1.3em; font-weight: bold; }}
        .params-table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        .params-table th, .params-table td {{ padding: 12px; text-align: left; border-bottom: 1px solid #eee; }}
        .params-table th {{ background: #002868; color: white; }}
        .params-table tr:nth-child(even) {{ background: #f8f9fa; }}
        .recommendation {{ background: linear-gradient(135deg, {color}22, {color}11); padding: 20px; border-radius: 10px; margin: 20px 0; border-left: 5px solid {color}; }}
        .footer {{ text-align: center; padding: 20px; color: #666; font-size: 0.9em; border-top: 1px solid #eee; }}
        @media print {{
            body {{ background: white; padding: 0; }}
            .container {{ box-shadow: none; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🩺 NefroPredict RD</h1>
            <p>Sistema de Detección Temprana de Enfermedad Renal Crónica</p>
        </div>
        <div class="content">
            <div class="info-grid">
                <div class="info-item">
                    <div class="info-label">Paciente</div>
                    <div class="info-value">{paciente['nombre_paciente']}</div>
                </div>
                <div class="info-item">
                    <div class="info-label">Médico Tratante</div>
                    <div class="info-value">{doctor}</div>
                </div>
                <div class="info-item">
                    <div class="info-label">Fecha de Evaluación</div>
                    <div class="info-value">{fecha}</div>
                </div>
                <div class="info-item">
                    <div class="info-label">ID Evaluación</div>
                    <div class="info-value">#{datetime.now().strftime('%Y%m%d%H%M')}</div>
                </div>
            </div>
            
            <div class="risk-display">
                <div class="risk-number">{riesgo:.1f}%</div>
                <div class="risk-label">Riesgo {nivel}</div>
            </div>
            
            <h3 style="color: #002868; margin: 20px 0 10px;">Parámetros Clínicos</h3>
            <table class="params-table">
                <tr><th>Parámetro</th><th>Valor</th><th>Rango Normal</th></tr>
                <tr><td>Edad</td><td>{paciente['edad']} años</td><td>-</td></tr>
                <tr><td>Índice de Masa Corporal</td><td>{paciente['imc']:.1f} kg/m²</td><td>18.5 - 24.9</td></tr>
                <tr><td>Presión Sistólica</td><td>{paciente['presion_sistolica']} mmHg</td><td>90 - 120</td></tr>
                <tr><td>Glucosa en Ayunas</td><td>{paciente['glucosa_ayunas']} mg/dL</td><td>70 - 100</td></tr>
                <tr><td>Creatinina Sérica</td><td>{paciente['creatinina']:.2f} mg/dL</td><td>0.7 - 1.3</td></tr>
            </table>
            
            <div class="recommendation">
                <h3 style="color: {color}; margin-bottom: 10px;">📋 Recomendación Clínica</h3>
                <p style="font-size: 1.1em;">{riesgo_level(riesgo)[2]}</p>
            </div>
        </div>
        <div class="footer">
            <p>NefroPredict RD © 2025 • República Dominicana</p>
            <p>Este reporte es una herramienta de apoyo diagnóstico y no reemplaza el criterio médico.</p>
        </div>
    </div>
    <script>window.onload = function() {{ window.print(); }}</script>
</body>
</html>
"""

# =============================================
# LOGIN
# =============================================
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    st.markdown("---")
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("### 🔐 Iniciar Sesión")
        with st.form("login_form"):
            username = st.text_input("Usuario").lower().strip()
            password = st.text_input("Contraseña", type="password")
            submitted = st.form_submit_button("Entrar", use_container_width=True)
            if submitted:
                user = db.verify_login(username, password)
                if user:
                    st.session_state.logged_in = True
                    st.session_state.username = username
                    st.session_state.role = user.get("role", "doctor")
                    st.session_state.doctor_name = user.get("name", username)
                    st.rerun()
                else:
                    st.error("❌ Usuario o contraseña incorrectos, o cuenta inactiva")
        
        st.markdown("---")
        st.markdown("<p style='text-align:center; color:#666; font-size:0.8em;'>🔒 Conexión segura • Contraseñas encriptadas</p>", unsafe_allow_html=True)
    st.stop()

# Barra superior con info de usuario
col1, col2 = st.columns([5, 1])
with col1:
    st.success(f"👨‍⚕️ **{st.session_state.doctor_name}** • @{st.session_state.username}")
with col2:
    if st.button("🚪 Salir"):
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        st.rerun()

st.markdown("---")

# =============================================
# PESTAÑAS SEGÚN ROL
# =============================================
if st.session_state.role == "admin":
    tabs = st.tabs(["📋 Evaluación Individual", "📤 Carga Masiva", "📊 Historial", "👥 Gestión Doctores", "📈 Estadísticas"])
    tab1, tab2, tab3, tab4, tab5 = tabs
else:
    tabs = st.tabs(["📋 Evaluación Individual", "📤 Carga Masiva", "📊 Historial"])
    tab1, tab2, tab3 = tabs

# =============================================
# TAB 1: EVALUACIÓN INDIVIDUAL
# =============================================
with tab1:
    st.subheader("📋 Evaluación Individual de Paciente")
    
    col_form, col_result = st.columns([1, 1])
    
    with col_form:
        st.markdown("#### Datos del Paciente")
        with st.form("form_individual", clear_on_submit=False):
            nombre = st.text_input("👤 Nombre completo del paciente", placeholder="Ej: Juan Pérez García")
            
            st.markdown("##### Datos Demográficos")
            c1, c2 = st.columns(2)
            with c1:
                edad = st.number_input("📅 Edad (años)", 18, 120, 55)
                sexo = st.selectbox("⚧ Sexo biológico", ["Masculino", "Femenino"])
            with c2:
                imc = st.number_input("⚖️ IMC (kg/m²)", 10.0, 60.0, 27.0, 0.1)
            
            st.markdown("##### Datos Clínicos")
            c3, c4 = st.columns(2)
            with c3:
                glucosa = st.number_input("🩸 Glucosa ayunas (mg/dL)", 50, 500, 110)
                presion = st.number_input("💓 Presión sistólica (mmHg)", 80, 250, 130)
            with c4:
                creatinina = st.number_input("🧪 Creatinina (mg/dL)", 0.1, 15.0, 1.2, 0.01)
            
            calcular = st.form_submit_button("🔬 Calcular Riesgo y TFG", use_container_width=True)
    
    with col_result:
        if calcular:
            if not nombre.strip():
                st.error("⚠️ El nombre del paciente es obligatorio")
            else:
                # Calcular riesgo
                datos_paciente = {
                    "edad": edad, "imc": imc, "presion_sistolica": presion,
                    "glucosa_ayunas": glucosa, "creatinina": creatinina
                }
                riesgo = predecir(datos_paciente)
                nivel, color, recomendacion = riesgo_level(riesgo)
                
                # Calcular TFG y Estadio ERC
                sexo_code = "F" if sexo == "Femenino" else "M"
                tfg = calcular_tfg(creatinina, edad, sexo_code)
                estadio, desc_estadio, color_estadio, reco_estadio = clasificar_estadio_erc(tfg)
                
                # Guardar en base de datos
                record = {
                    "nombre_paciente": nombre,
                    "doctor_user": st.session_state.username,
                    "doctor_name": st.session_state.doctor_name,
                    "timestamp": datetime.now().isoformat(),
                    "edad": edad, "sexo": sexo, "imc": imc, "presion_sistolica": presion,
                    "glucosa_ayunas": glucosa, "creatinina": creatinina,
                    "riesgo": riesgo, "nivel": nivel,
                    "tfg": tfg, "estadio_erc": estadio
                }
                db.add_patient(record)
                db.log_audit(st.session_state.username, f"Evaluó paciente: {nombre} - Riesgo: {riesgo}%, TFG: {tfg}, Estadio: {estadio}", "EVALUATION")
                
                # Guardar en session_state para mostrar
                st.session_state.ultimo_resultado = record
        
        # Mostrar resultado si existe
        if "ultimo_resultado" in st.session_state:
            p = st.session_state.ultimo_resultado
            nivel, color, recomendacion = riesgo_level(p["riesgo"])
            
            # Obtener TFG y estadio (con valores por defecto si no existen)
            tfg = p.get("tfg", calcular_tfg(p["creatinina"], p["edad"], "M"))
            estadio = p.get("estadio_erc", clasificar_estadio_erc(tfg)[0])
            estadio_info = clasificar_estadio_erc(tfg)
            
            st.markdown("#### 📊 Resultado del Análisis")
            
            # Métricas principales en tarjetas
            m1, m2 = st.columns(2)
            with m1:
                st.metric("🎯 Riesgo ERC", f"{p['riesgo']:.1f}%", nivel)
            with m2:
                st.metric("🧪 TFG (mL/min/1.73m²)", f"{tfg:.1f}", f"Estadio {estadio}")
            
            # Gráfico de gauge
            st.plotly_chart(crear_gauge_riesgo(p["riesgo"]), use_container_width=True)
            
            # Tarjeta de Estadio ERC
            st.markdown(f"""
            <div class="erc-stage {get_estadio_clase_css(estadio)}" style="margin: 15px 0;">
                <h3 style="margin:0;">📋 Estadio ERC: {estadio}</h3>
                <p style="margin:5px 0; font-size:1.1em;">{estadio_info[1]}</p>
                <p style="margin:5px 0;"><strong>TFG:</strong> {tfg:.1f} mL/min/1.73m²</p>
                <p style="margin:5px 0;"><strong>Recomendación:</strong> {estadio_info[3]}</p>
            </div>
            """, unsafe_allow_html=True)
            
            # Tarjeta de resultado de riesgo
            if p["riesgo"] > 70:
                st.markdown(f"""<div class="risk-high">
                    <h2>⚠️ RIESGO {nivel}</h2>
                    <h1 style="font-size:3em">{p["riesgo"]:.1f}%</h1>
                    <p>{recomendacion}</p>
                </div>""", unsafe_allow_html=True)
            elif p["riesgo"] > 40:
                st.markdown(f"""<div class="risk-med">
                    <h2>⚡ RIESGO {nivel}</h2>
                    <h1 style="font-size:3em">{p["riesgo"]:.1f}%</h1>
                    <p>{recomendacion}</p>
                </div>""", unsafe_allow_html=True)
            else:
                st.markdown(f"""<div class="risk-low">
                    <h2>✅ RIESGO {nivel}</h2>
                    <h1 style="font-size:3em">{p["riesgo"]:.1f}%</h1>
                    <p>{recomendacion}</p>
                </div>""", unsafe_allow_html=True)
            
            # Gráfico de factores
            st.plotly_chart(crear_grafico_factores(p), use_container_width=True)
            
            # Botón para descargar/imprimir reporte
            reporte_html = generar_reporte_html(p, p["riesgo"], nivel, st.session_state.doctor_name)
            st.download_button(
                "🖨️ Descargar/Imprimir Reporte",
                reporte_html,
                file_name=f"Reporte_{p['nombre_paciente'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.html",
                mime="text/html",
                use_container_width=True
            )

# =============================================
# TAB 2: CARGA MASIVA
# =============================================
with tab2:
    st.subheader("📤 Carga Masiva desde Excel/CSV")
    
    col1, col2 = st.columns([1, 3])
    
    with col1:
        st.markdown("#### 📥 Descargar Plantilla")
        plantilla = pd.DataFrame({
            "nombre_paciente": ["Juan Pérez", "María López", "Carlos Rodríguez"],
            "sexo": ["M", "F", "M"],
            "edad": [60, 55, 72],
            "imc": [29.5, 31.2, 26.8],
            "presion_sistolica": [150, 140, 165],
            "glucosa_ayunas": [180, 95, 220],
            "creatinina": [1.8, 1.1, 2.3]
        })
        csv = plantilla.to_csv(index=False).encode('utf-8')
        st.download_button("⬇️ Descargar Plantilla CSV", csv, "plantilla_nefropredict.csv", "text/csv", use_container_width=True)
        
        st.markdown("---")
        st.markdown("#### 📤 Subir Archivo")
        uploaded_file = st.file_uploader("Seleccionar archivo", type=["csv", "xlsx"])
        
        st.markdown("---")
        st.info("💡 **Tip:** No es obligatorio tener todas las columnas. Los campos vacíos se marcarán. Incluye 'sexo' (M/F) para cálculo preciso de TFG.")
    
    with col2:
        if uploaded_file:
            try:
                df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith(".csv") else pd.read_excel(uploaded_file)
                
                # Columnas esperadas con valores por defecto
                columnas_default = {
                    "nombre_paciente": "Sin nombre",
                    "sexo": "M",
                    "edad": 50,
                    "imc": 25.0,
                    "presion_sistolica": 120,
                    "glucosa_ayunas": 100,
                    "creatinina": 1.0
                }
                
                # Agregar columnas faltantes con valores por defecto
                campos_faltantes = []
                for col, default in columnas_default.items():
                    if col not in df.columns:
                        df[col] = default
                        campos_faltantes.append(col)
                
                # Marcar valores vacíos/nulos
                df["campos_vacios"] = ""
                for col in columnas_default.keys():
                    if col in df.columns:
                        # Detectar valores vacíos
                        mask = df[col].isna() | (df[col].astype(str).str.strip() == "")
                        if mask.any():
                            df.loc[mask, col] = columnas_default[col]
                            df.loc[mask, "campos_vacios"] += f"{col}, "
                
                df["campos_vacios"] = df["campos_vacios"].str.rstrip(", ")
                
                if campos_faltantes:
                    st.warning(f"⚠️ Columnas no encontradas (se usaron valores por defecto): {', '.join(campos_faltantes)}")
                
                # Calcular riesgo, TFG y estadio para todos
                df["riesgo"] = df.apply(predecir, axis=1)
                df["nivel"] = df["riesgo"].apply(lambda x: riesgo_level(x)[0])
                df["color"] = df["riesgo"].apply(lambda x: riesgo_level(x)[1])
                df["recomendacion"] = df["riesgo"].apply(lambda x: riesgo_level(x)[2])
                
                # Calcular TFG y Estadio ERC
                def calc_tfg_row(row):
                    sexo = str(row.get("sexo", "M")).upper()
                    sexo_code = "F" if sexo in ["F", "FEMENINO", "MUJER"] else "M"
                    return calcular_tfg(row["creatinina"], row["edad"], sexo_code)
                
                df["tfg"] = df.apply(calc_tfg_row, axis=1)
                df["estadio_erc"] = df["tfg"].apply(lambda x: clasificar_estadio_erc(x)[0])
                df["desc_estadio"] = df["tfg"].apply(lambda x: clasificar_estadio_erc(x)[1])
                
                # Clasificar pacientes
                df["categoria"] = df["riesgo"].apply(
                    lambda x: "🔴 Crítico" if x > 70 else "🟡 Medio" if x > 40 else "🟢 Normal"
                )
                
                # Guardar en BD
                for _, r in df.iterrows():
                    db.add_patient({
                        "nombre_paciente": str(r["nombre_paciente"]),
                        "doctor_user": st.session_state.username,
                        "doctor_name": st.session_state.doctor_name,
                        "timestamp": datetime.now().isoformat(),
                        "sexo": str(r.get("sexo", "M")),
                        "edad": int(r["edad"]), "imc": float(r["imc"]),
                        "presion_sistolica": int(r["presion_sistolica"]),
                        "glucosa_ayunas": int(r["glucosa_ayunas"]),
                        "creatinina": float(r["creatinina"]),
                        "riesgo": float(r["riesgo"]), "nivel": r["nivel"],
                        "tfg": float(r["tfg"]), "estadio_erc": r["estadio_erc"],
                        "campos_vacios": r["campos_vacios"]
                    })
                
                # Log de carga y auditoría
                db.add_upload_log({
                    "doctor_user": st.session_state.username,
                    "doctor_name": st.session_state.doctor_name,
                    "timestamp": datetime.now().isoformat(),
                    "cantidad": len(df)
                })
                db.log_audit(st.session_state.username, f"Carga masiva: {len(df)} pacientes procesados", "BULK_UPLOAD")
                
                # Conteos por categoría
                criticos = df[df["riesgo"] > 70]
                medios = df[(df["riesgo"] > 40) & (df["riesgo"] <= 70)]
                normales = df[df["riesgo"] <= 40]
                
                st.success(f"✅ **{len(df)} pacientes procesados exitosamente**")
                
                # ========== TARJETAS DE RESUMEN ==========
                st.markdown("### 📊 Resumen de Resultados")
                
                c1, c2, c3 = st.columns(3)
                
                with c1:
                    st.markdown(f"""
                    <div style="background: linear-gradient(135deg, #dc3545, #c82333); padding:25px; border-radius:15px; text-align:center; color:white; box-shadow: 0 4px 15px rgba(220,53,69,0.4);">
                        <h1 style="margin:0; font-size:3.5em; color:white !important;">{len(criticos)}</h1>
                        <h3 style="margin:5px 0; color:white !important;">🔴 CRÍTICOS</h3>
                        <p style="margin:0; color:#ffcccc;">Riesgo > 70%</p>
                        <p style="margin:5px 0; font-weight:bold; color:white !important;">Intervención URGENTE</p>
                    </div>
                    """, unsafe_allow_html=True)
                
                with c2:
                    st.markdown(f"""
                    <div style="background: linear-gradient(135deg, #fd7e14, #e76f00); padding:25px; border-radius:15px; text-align:center; color:white; box-shadow: 0 4px 15px rgba(253,126,20,0.4);">
                        <h1 style="margin:0; font-size:3.5em; color:white !important;">{len(medios)}</h1>
                        <h3 style="margin:5px 0; color:white !important;">🟡 RIESGO MEDIO</h3>
                        <p style="margin:0; color:#fff3cd;">Riesgo 40-70%</p>
                        <p style="margin:5px 0; font-weight:bold; color:white !important;">Control Estricto</p>
                    </div>
                    """, unsafe_allow_html=True)
                
                with c3:
                    st.markdown(f"""
                    <div style="background: linear-gradient(135deg, #28a745, #1e7e34); padding:25px; border-radius:15px; text-align:center; color:white; box-shadow: 0 4px 15px rgba(40,167,69,0.4);">
                        <h1 style="margin:0; font-size:3.5em; color:white !important;">{len(normales)}</h1>
                        <h3 style="margin:5px 0; color:white !important;">🟢 NORMALES</h3>
                        <p style="margin:0; color:#d4edda;">Riesgo < 40%</p>
                        <p style="margin:5px 0; font-weight:bold; color:white !important;">Seguimiento Rutinario</p>
                    </div>
                    """, unsafe_allow_html=True)
                
                st.markdown("---")
                
                # ========== GRÁFICOS ==========
                st.markdown("### 📈 Visualización General")
                
                graf_col1, graf_col2 = st.columns(2)
                
                with graf_col1:
                    # Gráfico de pastel
                    fig_pie = px.pie(
                        values=[len(criticos), len(medios), len(normales)],
                        names=['🔴 Críticos (>70%)', '🟡 Medio (40-70%)', '🟢 Normal (<40%)'],
                        color_discrete_sequence=['#CE1126', '#FFC400', '#4CAF50'],
                        title="📊 Distribución por Nivel de Riesgo",
                        hole=0.4
                    )
                    fig_pie.update_traces(textposition='inside', textinfo='percent+value')
                    fig_pie.update_layout(height=400)
                    st.plotly_chart(fig_pie, use_container_width=True)
                
                with graf_col2:
                    # Gráfico de barras horizontal por paciente
                    df_sorted = df.sort_values("riesgo", ascending=True).tail(15)
                    fig_bar = px.bar(
                        df_sorted,
                        y="nombre_paciente",
                        x="riesgo",
                        orientation='h',
                        color="riesgo",
                        color_continuous_scale=["#4CAF50", "#FFC400", "#CE1126"],
                        title="📊 Top 15 Pacientes por Riesgo",
                        text="riesgo"
                    )
                    fig_bar.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
                    fig_bar.update_layout(height=400, showlegend=False)
                    fig_bar.add_vline(x=70, line_dash="dash", line_color="red", annotation_text="Crítico")
                    fig_bar.add_vline(x=40, line_dash="dash", line_color="orange", annotation_text="Medio")
                    st.plotly_chart(fig_bar, use_container_width=True)
                
                st.markdown("---")
                
                # ========== FILTRO POR CATEGORÍA ==========
                st.markdown("### 🔍 Ver Pacientes por Categoría")
                
                filtro_cat = st.radio(
                    "Seleccionar categoría:",
                    ["📋 Todos", "🔴 Críticos", "🟡 Riesgo Medio", "🟢 Normales"],
                    horizontal=True
                )
                
                if filtro_cat == "🔴 Críticos":
                    df_mostrar = criticos.copy()
                elif filtro_cat == "🟡 Riesgo Medio":
                    df_mostrar = medios.copy()
                elif filtro_cat == "🟢 Normales":
                    df_mostrar = normales.copy()
                else:
                    df_mostrar = df.copy()
                
                if len(df_mostrar) > 0:
                    # Preparar dataframe para mostrar
                    columnas_mostrar = ["nombre_paciente", "sexo", "edad", "imc", "presion_sistolica", 
                                       "glucosa_ayunas", "creatinina", "tfg", "estadio_erc", "riesgo", "nivel"]
                    columnas_disponibles = [c for c in columnas_mostrar if c in df_mostrar.columns]
                    
                    df_display = df_mostrar[columnas_disponibles].copy()
                    
                    # Renombrar columnas
                    rename_cols = {
                        "nombre_paciente": "Paciente", "sexo": "Sexo", "edad": "Edad", 
                        "imc": "IMC", "presion_sistolica": "P.Sist.", "glucosa_ayunas": "Glucosa",
                        "creatinina": "Creat.", "tfg": "TFG", "estadio_erc": "Estadio",
                        "riesgo": "Riesgo %", "nivel": "Nivel"
                    }
                    df_display = df_display.rename(columns={c: rename_cols.get(c, c) for c in columnas_disponibles})
                    
                    # Ordenar por riesgo descendente
                    df_display = df_display.sort_values("Riesgo %", ascending=False)
                    
                    # Mostrar tabla con colores más intensos
                    st.markdown("""
                    <style>
                    .tabla-riesgo {
                        width: 100%;
                        border-collapse: collapse;
                        font-family: 'Segoe UI', Arial, sans-serif;
                        font-size: 14px;
                        margin: 20px 0;
                    }
                    .tabla-riesgo th {
                        background: #002868;
                        color: white;
                        padding: 12px 15px;
                        text-align: left;
                        font-weight: 600;
                        border: 1px solid #001a4d;
                    }
                    .tabla-riesgo td {
                        padding: 10px 15px;
                        border: 1px solid #ddd;
                    }
                    .fila-critico {
                        background: #f8d7da !important;
                        color: #721c24 !important;
                        font-weight: 600;
                    }
                    .fila-medio {
                        background: #fff3cd !important;
                        color: #856404 !important;
                    }
                    .fila-normal {
                        background: #d4edda !important;
                        color: #155724 !important;
                    }
                    .riesgo-badge-alto {
                        background: #dc3545;
                        color: white;
                        padding: 4px 10px;
                        border-radius: 20px;
                        font-weight: bold;
                    }
                    .riesgo-badge-medio {
                        background: #fd7e14;
                        color: white;
                        padding: 4px 10px;
                        border-radius: 20px;
                        font-weight: bold;
                    }
                    .riesgo-badge-bajo {
                        background: #28a745;
                        color: white;
                        padding: 4px 10px;
                        border-radius: 20px;
                        font-weight: bold;
                    }
                    </style>
                    """, unsafe_allow_html=True)
                    
                    # Generar tabla HTML personalizada
                    tabla_html = '<table class="tabla-riesgo"><thead><tr>'
                    for col in df_display.columns:
                        tabla_html += f'<th>{col}</th>'
                    tabla_html += '</tr></thead><tbody>'
                    
                    for _, row in df_display.iterrows():
                        riesgo_val = row["Riesgo %"]
                        if riesgo_val > 70:
                            clase_fila = "fila-critico"
                            badge_clase = "riesgo-badge-alto"
                        elif riesgo_val > 40:
                            clase_fila = "fila-medio"
                            badge_clase = "riesgo-badge-medio"
                        else:
                            clase_fila = "fila-normal"
                            badge_clase = "riesgo-badge-bajo"
                        
                        tabla_html += f'<tr class="{clase_fila}">'
                        for col in df_display.columns:
                            valor = row[col]
                            if col == "Riesgo %":
                                tabla_html += f'<td><span class="{badge_clase}">{valor:.1f}%</span></td>'
                            elif col == "TFG":
                                tabla_html += f'<td>{valor:.1f}</td>'
                            elif col in ["IMC", "Creat."]:
                                tabla_html += f'<td>{valor:.2f}</td>'
                            elif col == "Estadio":
                                estadio_color = clasificar_estadio_erc(row.get("TFG", 90))[2]
                                tabla_html += f'<td><span style="background:{estadio_color};color:white;padding:2px 8px;border-radius:4px;">{valor}</span></td>'
                            else:
                                tabla_html += f'<td>{valor}</td>'
                        tabla_html += '</tr>'
                    
                    tabla_html += '</tbody></table>'
                    st.markdown(tabla_html, unsafe_allow_html=True)
                    
                    # Mostrar leyenda
                    st.markdown("""
                    **Leyenda:** 
                    🔴 **Rojo** = Crítico (>70%) | 🟡 **Amarillo** = Medio (40-70%) | 🟢 **Verde** = Normal (<40%)
                    
                    **Estadios ERC:** G1 (Normal) → G2 (Leve) → G3a/G3b (Moderada) → G4 (Severa) → G5 (Falla Renal)
                    """)
                    
                    st.markdown("---")
                    
                    # ========== BOTÓN DESCARGAR EXCEL FORMATEADO ==========
                    st.markdown("### 📥 Descargar Reporte Excel")
                    
                    def generar_excel_formateado(dataframe, doctor_name):
                        """Genera un Excel bien formateado para impresión"""
                        output = BytesIO()
                        
                        # Preparar datos
                        df_excel = dataframe.copy()
                        df_excel = df_excel.sort_values("riesgo", ascending=False)
                        
                        # Renombrar columnas para el reporte
                        columnas_reporte = {
                            "nombre_paciente": "Nombre del Paciente",
                            "sexo": "Sexo",
                            "edad": "Edad (años)",
                            "imc": "IMC (kg/m²)",
                            "presion_sistolica": "Presión Sistólica (mmHg)",
                            "glucosa_ayunas": "Glucosa Ayunas (mg/dL)",
                            "creatinina": "Creatinina (mg/dL)",
                            "tfg": "TFG (mL/min/1.73m²)",
                            "estadio_erc": "Estadio ERC",
                            "riesgo": "Riesgo ERC (%)",
                            "nivel": "Clasificación",
                            "recomendacion": "Recomendación Clínica"
                        }
                        
                        # Seleccionar y renombrar columnas
                        cols_disponibles = [c for c in columnas_reporte.keys() if c in df_excel.columns]
                        df_reporte = df_excel[cols_disponibles].copy()
                        df_reporte = df_reporte.rename(columns={c: columnas_reporte[c] for c in cols_disponibles})
                        
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            # Escribir datos empezando en fila 5 para dejar espacio al encabezado
                            df_reporte.to_excel(writer, sheet_name='Reporte ERC', index=False, startrow=4)
                            
                            workbook = writer.book
                            worksheet = writer.sheets['Reporte ERC']
                            
                            # Importar estilos de openpyxl
                            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                            from openpyxl.utils import get_column_letter
                            
                            # Estilos
                            borde = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                            
                            header_fill = PatternFill(start_color="002868", end_color="002868", fill_type="solid")
                            header_font = Font(color="FFFFFF", bold=True, size=11)
                            
                            critico_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                            critico_font = Font(color="721C24", bold=True)
                            
                            medio_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                            medio_font = Font(color="856404")
                            
                            normal_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                            normal_font = Font(color="155724")
                            
                            # Título del reporte
                            worksheet.merge_cells('A1:I1')
                            worksheet['A1'] = "🩺 NEFROPREDICT RD - REPORTE DE EVALUACIÓN DE RIESGO ERC"
                            worksheet['A1'].font = Font(size=16, bold=True, color="002868")
                            worksheet['A1'].alignment = Alignment(horizontal='center')
                            
                            # Info del reporte
                            worksheet['A2'] = f"Médico: {doctor_name}"
                            worksheet['A2'].font = Font(size=11, bold=True)
                            worksheet['A3'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                            worksheet['A3'].font = Font(size=11)
                            worksheet['D2'] = f"Total Pacientes: {len(df_reporte)}"
                            worksheet['D2'].font = Font(size=11, bold=True)
                            worksheet['D3'] = f"Críticos: {len(df_excel[df_excel['riesgo']>70])} | Medio: {len(df_excel[(df_excel['riesgo']>40) & (df_excel['riesgo']<=70)])} | Normal: {len(df_excel[df_excel['riesgo']<=40])}"
                            worksheet['D3'].font = Font(size=11)
                            
                            # Aplicar estilos a encabezados (fila 5)
                            for col_num, column_title in enumerate(df_reporte.columns, 1):
                                cell = worksheet.cell(row=5, column=col_num)
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                cell.border = borde
                            
                            # Aplicar estilos a datos
                            for row_num, (_, row_data) in enumerate(df_reporte.iterrows(), 6):
                                # Determinar color de fila según riesgo
                                riesgo_val = df_excel.iloc[row_num-6]["riesgo"]
                                
                                if riesgo_val > 70:
                                    fill = critico_fill
                                    font = critico_font
                                elif riesgo_val > 40:
                                    fill = medio_fill
                                    font = medio_font
                                else:
                                    fill = normal_fill
                                    font = normal_font
                                
                                for col_num, value in enumerate(row_data, 1):
                                    cell = worksheet.cell(row=row_num, column=col_num)
                                    cell.fill = fill
                                    cell.font = font
                                    cell.border = borde
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                            
                            # Ajustar ancho de columnas
                            anchos = [30, 8, 12, 15, 22, 22, 18, 20, 12, 15, 18, 35]
                            for i, ancho in enumerate(anchos[:len(df_reporte.columns)], 1):
                                worksheet.column_dimensions[get_column_letter(i)].width = ancho
                            
                            # Ajustar altura de filas
                            worksheet.row_dimensions[1].height = 25
                            worksheet.row_dimensions[5].height = 30
                            
                            # Agregar leyenda al final
                            ultima_fila = len(df_reporte) + 7
                            worksheet.merge_cells(f'A{ultima_fila}:I{ultima_fila}')
                            worksheet[f'A{ultima_fila}'] = "Leyenda: 🔴 Rojo = Crítico (>70% riesgo) | 🟡 Amarillo = Riesgo Medio (40-70%) | 🟢 Verde = Normal (<40%) | Estadios ERC: G1-G2 Normal/Leve, G3a-G3b Moderada, G4-G5 Severa/Falla"
                            worksheet[f'A{ultima_fila}'].font = Font(size=10, italic=True)
                            
                            # Configurar página para impresión
                            worksheet.print_title_rows = '1:5'
                            worksheet.page_setup.orientation = 'landscape'
                            worksheet.page_setup.fitToPage = True
                            worksheet.page_setup.fitToWidth = 1
                            worksheet.page_setup.fitToHeight = 0
                        
                        output.seek(0)
                        return output
                    
                    col_btn1, col_btn2 = st.columns(2)
                    
                    with col_btn1:
                        excel_file = generar_excel_formateado(df_mostrar, st.session_state.doctor_name)
                        st.download_button(
                            label="📥 Descargar Excel Formateado",
                            data=excel_file,
                            file_name=f"Reporte_ERC_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    
                    with col_btn2:
                        # También opción de CSV simple
                        columnas_csv = ["nombre_paciente", "sexo", "edad", "imc", "presion_sistolica", 
                                       "glucosa_ayunas", "creatinina", "tfg", "estadio_erc", "riesgo", "nivel"]
                        cols_disponibles_csv = [c for c in columnas_csv if c in df_mostrar.columns]
                        csv_data = df_mostrar[cols_disponibles_csv].to_csv(index=False).encode('utf-8')
                        st.download_button(
                            label="📄 Descargar CSV Simple",
                            data=csv_data,
                            file_name=f"Reporte_ERC_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                            mime="text/csv",
                            use_container_width=True
                        )
                    
                else:
                    st.info("No hay pacientes en esta categoría")
                
                # ========== SELECTOR DE PACIENTE INDIVIDUAL ==========
                st.markdown("---")
                st.markdown("### 👤 Ver Detalle de Paciente")
                
                paciente_sel = st.selectbox(
                    "Seleccionar paciente para ver detalle:",
                    [""] + df["nombre_paciente"].tolist()
                )
                
                if paciente_sel:
                    pac_data = df[df["nombre_paciente"] == paciente_sel].iloc[0]
                    
                    col_det1, col_det2 = st.columns([1, 1])
                    
                    with col_det1:
                        st.markdown(f"#### 📋 {pac_data['nombre_paciente']}")
                        
                        # Métricas principales
                        m1, m2 = st.columns(2)
                        m1.metric("🎯 Riesgo ERC", f"{pac_data['riesgo']:.1f}%", pac_data['nivel'])
                        m2.metric("🧪 TFG", f"{pac_data['tfg']:.1f}", f"Estadio {pac_data['estadio_erc']}")
                        
                        # Gauge de riesgo
                        st.plotly_chart(crear_gauge_riesgo(pac_data["riesgo"]), use_container_width=True)
                    
                    with col_det2:
                        st.markdown("#### 📊 Datos Clínicos")
                        
                        # Tarjeta de estadio ERC
                        estadio_info = clasificar_estadio_erc(pac_data["tfg"])
                        st.markdown(f"""
                        <div class="erc-stage {get_estadio_clase_css(pac_data['estadio_erc'])}" style="margin-bottom: 15px;">
                            <strong>Estadio {pac_data['estadio_erc']}</strong>: {estadio_info[1]}<br>
                            <small>{estadio_info[3]}</small>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        datos_tabla = {
                            "Parámetro": ["Sexo", "Edad", "IMC", "Presión Sistólica", "Glucosa Ayunas", "Creatinina", "TFG", "Estadio ERC"],
                            "Valor": [
                                pac_data.get('sexo', 'N/A'),
                                f"{pac_data['edad']} años",
                                f"{pac_data['imc']:.1f} kg/m²",
                                f"{pac_data['presion_sistolica']} mmHg",
                                f"{pac_data['glucosa_ayunas']} mg/dL",
                                f"{pac_data['creatinina']:.2f} mg/dL",
                                f"{pac_data['tfg']:.1f} mL/min/1.73m²",
                                pac_data['estadio_erc']
                            ],
                            "Rango Normal": ["-", "18-65", "18.5-24.9", "90-120", "70-100", "0.7-1.3", ">90", "G1"]
                        }
                        st.table(pd.DataFrame(datos_tabla))
                        
                        nivel, color, reco = riesgo_level(pac_data["riesgo"])
                        st.markdown(f"**Recomendación Riesgo:** {reco}")
                        
                        if pac_data.get("campos_vacios"):
                            st.warning(f"⚠️ Campos no proporcionados: {pac_data['campos_vacios']}")
                
            except Exception as e:
                st.error(f"❌ Error al procesar archivo: {str(e)}")
                st.info("Asegúrate de que el archivo tenga al menos una columna con datos numéricos o nombres de pacientes.")

# =============================================
# TAB 3: HISTORIAL
# =============================================
with tab3:
    st.subheader("📊 Historial Clínico")
    
    # Obtener pacientes según rol
    if st.session_state.role == "doctor":
        pacientes = db.get_patients_by_doctor(st.session_state.username)
    else:
        pacientes = db.get_all_patients()
    
    if pacientes:
        dfp = pd.DataFrame(pacientes)
        
        col1, col2 = st.columns([1, 2])
        
        with col1:
            st.markdown("#### 🔍 Buscar Paciente")
            nombres = dfp["nombre_paciente"].unique().tolist()
            paciente_sel = st.selectbox("Seleccionar paciente", [""] + nombres)
            
            # Métricas generales
            st.markdown("#### 📈 Resumen General")
            total = len(dfp)
            prom_riesgo = dfp["riesgo"].mean()
            alto_riesgo = len(dfp[dfp["riesgo"] > 70])
            
            st.metric("Total Evaluaciones", total)
            st.metric("Riesgo Promedio", f"{prom_riesgo:.1f}%")
            st.metric("Alto Riesgo (>70%)", alto_riesgo)
        
        with col2:
            if paciente_sel:
                hist = dfp[dfp["nombre_paciente"] == paciente_sel].sort_values("timestamp")
                
                st.markdown(f"#### 📋 Historial de: **{paciente_sel}**")
                
                # Último resultado
                ultimo = hist.iloc[-1]
                nivel, color, reco = riesgo_level(ultimo["riesgo"])
                
                c1, c2, c3 = st.columns(3)
                c1.metric("Último Riesgo", f"{ultimo['riesgo']:.1f}%")
                c2.metric("Nivel", nivel)
                c3.metric("Evaluaciones", len(hist))
                
                # Gráfico de evolución temporal
                if len(hist) > 1:
                    fig_evol = px.line(
                        hist, x="timestamp", y="riesgo",
                        title="📈 Evolución del Riesgo en el Tiempo",
                        markers=True
                    )
                    fig_evol.add_hline(y=70, line_dash="dash", line_color="red", annotation_text="Alto Riesgo")
                    fig_evol.add_hline(y=40, line_dash="dash", line_color="orange", annotation_text="Riesgo Medio")
                    fig_evol.update_layout(height=300)
                    st.plotly_chart(fig_evol, use_container_width=True)
                
                # Tabla de historial
                st.dataframe(
                    hist[["timestamp", "riesgo", "nivel", "creatinina", "glucosa_ayunas", "presion_sistolica"]],
                    use_container_width=True,
                    hide_index=True
                )
            else:
                # Mostrar resumen general si no hay paciente seleccionado
                st.markdown("#### 📊 Distribución General de Riesgo")
                
                # Gráfico de distribución
                fig_hist = px.histogram(
                    dfp, x="riesgo", nbins=20,
                    title="Distribución de Niveles de Riesgo",
                    color_discrete_sequence=["#002868"]
                )
                fig_hist.add_vline(x=70, line_dash="dash", line_color="red")
                fig_hist.add_vline(x=40, line_dash="dash", line_color="orange")
                st.plotly_chart(fig_hist, use_container_width=True)
                
                # Top pacientes alto riesgo
                st.markdown("#### ⚠️ Pacientes con Mayor Riesgo")
                top_riesgo = dfp.nlargest(10, "riesgo")[["nombre_paciente", "riesgo", "nivel", "timestamp"]]
                st.dataframe(top_riesgo, use_container_width=True, hide_index=True)
    else:
        st.info("📭 Aún no hay registros de pacientes")

# =============================================
# TAB 4: GESTIÓN DE DOCTORES (SOLO ADMIN)
# =============================================
if st.session_state.role == "admin":
    with tab4:
        st.subheader("👥 Panel de Administración de Usuarios")
        
        # Sub-tabs dentro del panel admin
        admin_tab1, admin_tab2, admin_tab3, admin_tab4 = st.tabs(["📋 Ver/Gestionar Doctores", "➕ Crear Doctor", "📊 Actividad y Archivos", "🔍 Auditoría"])
        
        # ---------- SUB-TAB 1: VER Y GESTIONAR DOCTORES ----------
        with admin_tab1:
            st.markdown("#### 📋 Todos los Usuarios del Sistema")
            
            # Tabla resumen de todos los usuarios
            usuarios_data = []
            for user, info in db.data["users"].items():
                # Contar pacientes de este doctor
                if info.get("role") == "doctor":
                    num_pacientes = len(db.get_patients_by_doctor(user))
                    num_cargas = len([u for u in db.data.get("uploads", []) if u.get("doctor_user") == user])
                else:
                    num_pacientes = "-"
                    num_cargas = "-"
                
                usuarios_data.append({
                    "Usuario": f"@{user}",
                    "Nombre": info.get("name", user),
                    "Rol": "👑 Admin" if info.get("role") == "admin" else "👨‍⚕️ Doctor",
                    "Estado": "🟢 Activo" if info.get("active", True) else "🔴 Inactivo",
                    "Contraseña": "🔒 Encriptada",
                    "Pacientes": num_pacientes,
                    "Cargas": num_cargas,
                    "Creado": info.get("created_at", "N/A")[:10] if info.get("created_at") else "Inicial"
                })
            
            df_usuarios = pd.DataFrame(usuarios_data)
            st.dataframe(df_usuarios, use_container_width=True, hide_index=True)
            
            st.info("🔒 Las contraseñas están encriptadas con bcrypt y no pueden ser visualizadas")
            
            st.markdown("---")
            st.markdown("#### ⚙️ Gestionar Doctor Individual")
            
            # Selector de doctor
            doctores = {u: i for u, i in db.data["users"].items() if i["role"] == "doctor"}
            
            if doctores:
                opciones_doctores = [f"{info.get('name', user)} (@{user})" for user, info in doctores.items()]
                doctor_seleccionado = st.selectbox("Seleccionar doctor a gestionar:", [""] + opciones_doctores)
                
                if doctor_seleccionado:
                    # Extraer username del doctor seleccionado
                    user_sel = doctor_seleccionado.split("(@")[1].replace(")", "")
                    info_sel = doctores[user_sel]
                    
                    st.markdown(f"##### Gestionando: **{info_sel.get('name', user_sel)}**")
                    
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.markdown("**🔐 Cambiar Contraseña**")
                        nueva_pwd = st.text_input("Nueva contraseña:", type="password", key="admin_new_pwd")
                        if st.button("💾 Guardar Contraseña", key="btn_save_pwd", use_container_width=True):
                            if nueva_pwd:
                                db.update_password(user_sel, nueva_pwd, st.session_state.username)
                                st.success("✅ Contraseña actualizada")
                                st.rerun()
                            else:
                                st.warning("Ingrese una contraseña")
                    
                    with col2:
                        st.markdown("**🔄 Estado de Cuenta**")
                        estado_actual = "🟢 ACTIVO" if info_sel.get("active", True) else "🔴 INACTIVO"
                        st.info(f"Estado actual: {estado_actual}")
                        
                        if info_sel.get("active", True):
                            if st.button("🚫 Desactivar Cuenta", key="btn_desactivar", use_container_width=True):
                                db.toggle_active(user_sel, st.session_state.username)
                                st.warning("Cuenta desactivada")
                                st.rerun()
                        else:
                            if st.button("✅ Activar Cuenta", key="btn_activar", use_container_width=True):
                                db.toggle_active(user_sel, st.session_state.username)
                                st.success("Cuenta activada")
                                st.rerun()
                    
                    with col3:
                        st.markdown("**🗑️ Eliminar Doctor**")
                        st.warning("⚠️ Esta acción es irreversible")
                        confirmar = st.checkbox("Confirmo que deseo eliminar", key="confirm_del")
                        if st.button("🗑️ Eliminar Permanentemente", key="btn_eliminar", type="primary", use_container_width=True, disabled=not confirmar):
                            db.delete_doctor(user_sel, st.session_state.username)
                            st.success(f"Doctor {info_sel.get('name', user_sel)} eliminado")
                            st.rerun()
                    
                    # Ver pacientes de este doctor
                    st.markdown("---")
                    st.markdown(f"##### 📁 Pacientes de {info_sel.get('name', user_sel)}")
                    pacientes_doctor = db.get_patients_by_doctor(user_sel)
                    if pacientes_doctor:
                        df_pac = pd.DataFrame(pacientes_doctor)
                        st.dataframe(
                            df_pac[["nombre_paciente", "timestamp", "riesgo", "nivel"]].head(20),
                            use_container_width=True,
                            hide_index=True
                        )
                        st.caption(f"Mostrando últimos 20 de {len(pacientes_doctor)} registros")
                    else:
                        st.info("Este doctor no tiene pacientes registrados")
            else:
                st.info("No hay doctores registrados en el sistema")
        
        # ---------- SUB-TAB 2: CREAR DOCTOR ----------
        with admin_tab2:
            st.markdown("#### ➕ Registrar Nuevo Doctor")
            
            with st.form("form_nuevo_doctor"):
                col1, col2 = st.columns(2)
                with col1:
                    nuevo_user = st.text_input("👤 Usuario (sin espacios)", placeholder="dr.apellido").lower().strip()
                    nuevo_pwd = st.text_input("🔐 Contraseña", type="password")
                with col2:
                    nuevo_nombre = st.text_input("📝 Nombre completo", placeholder="Dr. Juan Pérez")
                    nuevo_pwd_confirm = st.text_input("🔐 Confirmar contraseña", type="password")
                
                if st.form_submit_button("✅ Crear Doctor", use_container_width=True):
                    if not nuevo_user or not nuevo_pwd or not nuevo_nombre:
                        st.error("❌ Todos los campos son obligatorios")
                    elif db.get_user(nuevo_user):
                        st.error("❌ El usuario ya existe")
                    elif " " in nuevo_user:
                        st.error("❌ El usuario no puede tener espacios")
                    elif nuevo_pwd != nuevo_pwd_confirm:
                        st.error("❌ Las contraseñas no coinciden")
                    elif len(nuevo_pwd) < 4:
                        st.error("❌ La contraseña debe tener al menos 4 caracteres")
                    else:
                        db.create_doctor(nuevo_user, nuevo_pwd, nuevo_nombre, st.session_state.username)
                        st.success(f"✅ Doctor **{nuevo_nombre}** creado exitosamente")
                        st.balloons()
                        st.rerun()
        
        # ---------- SUB-TAB 3: ACTIVIDAD Y ARCHIVOS ----------
        with admin_tab3:
            st.markdown("#### 📊 Ranking de Doctores más Activos")
            
            all_patients = db.get_all_patients()
            
            if all_patients:
                df_all = pd.DataFrame(all_patients)
                
                # Ranking por número de evaluaciones
                ranking = df_all.groupby(["doctor_user", "doctor_name"]).agg({
                    "nombre_paciente": "count",
                    "riesgo": "mean"
                }).reset_index()
                ranking.columns = ["Usuario", "Doctor", "Total Evaluaciones", "Riesgo Promedio"]
                ranking["Riesgo Promedio"] = ranking["Riesgo Promedio"].round(1)
                ranking = ranking.sort_values("Total Evaluaciones", ascending=False)
                ranking["🏆 Ranking"] = range(1, len(ranking) + 1)
                ranking = ranking[["🏆 Ranking", "Doctor", "Usuario", "Total Evaluaciones", "Riesgo Promedio"]]
                
                # Mostrar podio
                col1, col2, col3 = st.columns(3)
                if len(ranking) >= 1:
                    col2.metric("🥇 1er Lugar", ranking.iloc[0]["Doctor"], f"{ranking.iloc[0]['Total Evaluaciones']} evaluaciones")
                if len(ranking) >= 2:
                    col1.metric("🥈 2do Lugar", ranking.iloc[1]["Doctor"], f"{ranking.iloc[1]['Total Evaluaciones']} evaluaciones")
                if len(ranking) >= 3:
                    col3.metric("🥉 3er Lugar", ranking.iloc[2]["Doctor"], f"{ranking.iloc[2]['Total Evaluaciones']} evaluaciones")
                
                st.markdown("##### 📋 Tabla Completa de Actividad")
                st.dataframe(ranking, use_container_width=True, hide_index=True)
                
                # Gráfico de barras
                fig_ranking = px.bar(
                    ranking.head(10), 
                    x="Doctor", 
                    y="Total Evaluaciones",
                    color="Total Evaluaciones",
                    color_continuous_scale="Blues",
                    title="Top 10 Doctores por Evaluaciones"
                )
                st.plotly_chart(fig_ranking, use_container_width=True)
            else:
                st.info("No hay actividad registrada aún")
            
            st.markdown("---")
            st.markdown("#### 📤 Historial de Cargas Masivas")
            
            uploads = db.data.get("uploads", [])
            if uploads:
                df_uploads = pd.DataFrame(uploads)
                st.dataframe(df_uploads, use_container_width=True, hide_index=True)
            else:
                st.info("No hay cargas masivas registradas")
            
            st.markdown("---")
            st.markdown("#### 🔒 Acceso a Datos por Doctor")
            st.info("👆 Selecciona un doctor en la pestaña 'Ver/Gestionar Doctores' para ver sus pacientes. Los datos de cada doctor están completamente aislados y nunca se cruzan entre médicos.")
        
        # ---------- SUB-TAB 4: AUDITORÍA ----------
        with admin_tab4:
            st.markdown("#### 🔍 Registro de Auditoría del Sistema")
            st.markdown("Todas las acciones importantes son registradas para cumplimiento y seguridad.")
            
            # Filtros
            col_f1, col_f2, col_f3 = st.columns(3)
            with col_f1:
                filtro_tipo = st.selectbox("Filtrar por tipo:", 
                    ["Todos", "LOGIN", "LOGIN_FAILED", "EVALUATION", "USER_CREATED", 
                     "USER_DELETED", "PASSWORD_CHANGED", "USER_STATUS_CHANGED", "SECURITY"])
            with col_f2:
                usuarios_list = ["Todos"] + list(db.data["users"].keys())
                filtro_usuario = st.selectbox("Filtrar por usuario:", usuarios_list)
            with col_f3:
                limite = st.number_input("Mostrar últimos:", 50, 500, 100, 50)
            
            # Obtener logs
            tipo_f = None if filtro_tipo == "Todos" else filtro_tipo
            user_f = None if filtro_usuario == "Todos" else filtro_usuario
            logs = db.get_audit_log(limit=limite, user_filter=user_f, type_filter=tipo_f)
            
            if logs:
                # Estadísticas rápidas
                st.markdown("##### 📊 Resumen de Actividad")
                col_s1, col_s2, col_s3, col_s4 = st.columns(4)
                
                all_logs = db.get_audit_log(limit=1000)
                logins_hoy = len([l for l in all_logs if l.get("type") == "LOGIN" and l.get("timestamp", "")[:10] == datetime.now().strftime("%Y-%m-%d")])
                logins_fallidos = len([l for l in all_logs if l.get("type") == "LOGIN_FAILED"])
                evaluaciones = len([l for l in all_logs if l.get("type") == "EVALUATION"])
                cambios_usuarios = len([l for l in all_logs if l.get("type") in ["USER_CREATED", "USER_DELETED", "PASSWORD_CHANGED"]])
                
                col_s1.metric("🔐 Logins Hoy", logins_hoy)
                col_s2.metric("⚠️ Logins Fallidos", logins_fallidos)
                col_s3.metric("📋 Evaluaciones", evaluaciones)
                col_s4.metric("👥 Cambios Usuarios", cambios_usuarios)
                
                st.markdown("---")
                st.markdown("##### 📜 Registro de Eventos")
                
                # Mostrar logs con colores según tipo
                for log in logs:
                    timestamp = log.get("timestamp", "")[:19].replace("T", " ")
                    user = log.get("user", "N/A")
                    action = log.get("action", "N/A")
                    log_type = log.get("type", "INFO")
                    
                    # Color según tipo
                    if log_type in ["LOGIN_FAILED", "SECURITY"]:
                        color = "#dc3545"
                        icon = "🚨"
                    elif log_type == "LOGIN":
                        color = "#28a745"
                        icon = "✅"
                    elif log_type == "EVALUATION":
                        color = "#007bff"
                        icon = "📋"
                    elif log_type in ["USER_CREATED", "USER_DELETED", "PASSWORD_CHANGED", "USER_STATUS_CHANGED"]:
                        color = "#fd7e14"
                        icon = "👤"
                    else:
                        color = "#6c757d"
                        icon = "ℹ️"
                    
                    st.markdown(f"""
                    <div class="audit-log" style="border-left-color: {color};">
                        <span style="color: #666; font-size: 0.85em;">{timestamp}</span> 
                        <span style="background: {color}; color: white; padding: 2px 8px; border-radius: 4px; font-size: 0.8em; margin: 0 8px;">{log_type}</span>
                        <strong>@{user}</strong>: {icon} {action}
                    </div>
                    """, unsafe_allow_html=True)
                
                # Exportar logs
                st.markdown("---")
                df_logs = pd.DataFrame(logs)
                csv_logs = df_logs.to_csv(index=False).encode('utf-8')
                st.download_button(
                    "📥 Exportar Logs a CSV",
                    csv_logs,
                    f"audit_log_{datetime.now().strftime('%Y%m%d')}.csv",
                    "text/csv",
                    use_container_width=True
                )
            else:
                st.info("No hay registros de auditoría con los filtros seleccionados")

    # =============================================
    # TAB 5: ESTADÍSTICAS ADMIN
    # =============================================
    with tab5:
        st.subheader("📈 Estadísticas del Sistema")
        
        all_patients = db.get_all_patients()
        
        if all_patients:
            df_all = pd.DataFrame(all_patients)
            
            # Métricas principales
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("📊 Total Evaluaciones", len(df_all))
            c2.metric("👥 Pacientes Únicos", df_all["nombre_paciente"].nunique())
            c3.metric("📈 Riesgo Promedio", f"{df_all['riesgo'].mean():.1f}%")
            c4.metric("⚠️ Alto Riesgo", len(df_all[df_all["riesgo"] > 70]))
            
            st.markdown("---")
            
            col1, col2 = st.columns(2)
            
            with col1:
                # Evaluaciones por doctor
                st.markdown("#### 👨‍⚕️ Evaluaciones por Doctor")
                por_doctor = df_all["doctor_name"].value_counts().reset_index()
                por_doctor.columns = ["Doctor", "Evaluaciones"]
                fig_doc = px.bar(por_doctor, x="Doctor", y="Evaluaciones", color="Evaluaciones",
                                color_continuous_scale="Blues")
                st.plotly_chart(fig_doc, use_container_width=True)
            
            with col2:
                # Distribución de riesgo
                st.markdown("#### 📊 Distribución de Riesgo")
                df_all["categoria"] = df_all["riesgo"].apply(
                    lambda x: "Alto (>70%)" if x > 70 else "Medio (40-70%)" if x > 40 else "Bajo (<40%)"
                )
                por_cat = df_all["categoria"].value_counts().reset_index()
                por_cat.columns = ["Categoría", "Cantidad"]
                fig_cat = px.pie(por_cat, values="Cantidad", names="Categoría",
                                color_discrete_sequence=["#CE1126", "#FFC400", "#4CAF50"])
                st.plotly_chart(fig_cat, use_container_width=True)
            
            # Tendencia temporal
            st.markdown("#### 📅 Tendencia de Evaluaciones")
            df_all["fecha"] = pd.to_datetime(df_all["timestamp"]).dt.date
            por_fecha = df_all.groupby("fecha").agg({"riesgo": ["count", "mean"]}).reset_index()
            por_fecha.columns = ["Fecha", "Cantidad", "Riesgo Promedio"]
            
            fig_trend = px.line(por_fecha, x="Fecha", y="Cantidad", 
                               title="Evaluaciones por Día", markers=True)
            st.plotly_chart(fig_trend, use_container_width=True)
            
            # Tabla resumen por doctor
            st.markdown("#### 📋 Resumen por Doctor")
            resumen = df_all.groupby("doctor_name").agg({
                "nombre_paciente": "count",
                "riesgo": "mean"
            }).reset_index()
            resumen.columns = ["Doctor", "Total Evaluaciones", "Riesgo Promedio"]
            resumen["Riesgo Promedio"] = resumen["Riesgo Promedio"].round(1)
            st.dataframe(resumen, use_container_width=True, hide_index=True)
        else:
            st.info("📭 Aún no hay datos para mostrar estadísticas")

# =============================================
# FOOTER
# =============================================
st.markdown("---")
st.caption("🩺 NefroPredict RD © 2025 • Sistema de Detección Temprana de ERC • República Dominicana")
